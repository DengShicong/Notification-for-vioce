from pysnmp.hlapi import SnmpEngine, CommunityData, UdpTransportTarget, ContextData, ObjectType, ObjectIdentity, nextCmd
import pandas as pd
from datetime import datetime
import time
from openpyxl.styles import PatternFill
import pyttsx3
import logging
from concurrent.futures import ThreadPoolExecutor

def speak_text(text):
    engine = pyttsx3.init()
    engine.say(text)
    engine.runAndWait()

# 根据消息播放语音
def play_warning_message(message):
    speak_text(message)
def apply_color(writer, sheet_name, changes_df):
    """
    为指定工作表应用颜色

    参数：
    writer: 写入工作簿的ExcelWriter对象
    sheet_name: 工作表名称
    changes_df: 包含变更数据的DataFrame对象
    """

    # 获取工作簿和工作表
    workbook = writer.book
    worksheet = workbook[sheet_name]

    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    green_fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')

    # 遍历变更数据的行和列
    for r_idx, row in enumerate(changes_df.itertuples(), start=2):  # Excel 行号从1开始，标题占据第1行
        if row.ifOperStatus == '2':
            worksheet.cell(row=r_idx, column=changes_df.columns.get_loc('ifOperStatus') + 1).fill = red_fill
        elif row.ifOperStatus == '1':
            worksheet.cell(row=r_idx, column=changes_df.columns.get_loc('ifOperStatus') + 1).fill = green_fill


def snmp_walk(ip, community, oid):
    """
    使用SNMP遍历OID返回结果的函数

    参数：
    ip：字符串类型，设备的IP地址
    community：字符串类型，SNMP的community名称
    oid：字符串类型，要遍历的OID

    返回值：
    list类型，遍历结果

    """
    result = []
    for (errorIndication, errorStatus, errorIndex, varBinds) in nextCmd(
        SnmpEngine(),
        CommunityData(community),
        UdpTransportTarget((ip, 161)),
        ContextData(),
        ObjectType(ObjectIdentity(oid)),
        lexicographicMode=False):
        if errorIndication:
            print(f"故障设备IP地址 {ip}: {errorIndication}")
            return None  # 返回 None 表示有错误
        elif errorStatus:
            print('%s at %s' % (errorStatus.prettyPrint(), errorIndex and varBinds[int(errorIndex) - 1][0] or '?'))
            break
        else:
            for varBind in varBinds:
                result.append(varBind[1].prettyPrint())
    return result



def process_device(device_type, ip, oids, previous_data):
    print(f"正在处理 {device_type} 设备，IP地址为 {ip}")
    interface_info = {key: [] for key in oids.keys()}
    error_occurred = False
    for key, oid in oids.items():
        results = snmp_walk(ip, 'public', oid)
        if results is None:
            error_occurred = True
            break
        interface_info[key].extend(results)

    if error_occurred:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        error_filename = f"error_{ip}_{timestamp}.xlsx"
        df_error = pd.DataFrame({'Error': [f"IP {ip} 连接失败"]})
        df_error.to_excel(error_filename, index=False)
        print(f"错误信息已写入 {error_filename}")
    else:
        df = pd.DataFrame(interface_info)
        key = f"{device_type}_{ip}"
        if key in previous_data:
            if 'ifOperStatus' in df:
                changed_status = df[df['ifOperStatus'] != previous_data[key].get('ifOperStatus', pd.Series())]
                if not changed_status.empty:
                    for _, row in changed_status.iterrows():
                        if row['ifOperStatus'] == '1':
                            play_warning_message("端口已启用")
                        elif row['ifOperStatus'] == '2':
                            play_warning_message("端口已断开")

                    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                    filename = f"combined_info_{timestamp}.xlsx"
                    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                        changed_status.to_excel(writer, sheet_name=key, index=False)
                        apply_color(writer, key, changed_status)
                    print(f"数据已写入 {filename}")

        previous_data[key] = df

# 主逻辑
def main():
    oids_by_device_type = {
        'OmniSwitch': {
            "ifIndex": ".1.3.6.1.2.1.2.2.1.1",
            "ifDescr": ".1.3.6.1.2.1.2.2.1.2",
            "ifType": ".1.3.6.1.2.1.2.2.1.3",
            "ifMtu": ".1.3.6.1.2.1.2.2.1.4",
            "ifSpeed": ".1.3.6.1.2.1.2.2.1.5",
            "ifPhysAddress": ".1.3.6.1.2.1.2.2.1.6",
            "ifAdminStatus": ".1.3.6.1.2.1.2.2.1.7",
            "ifOperStatus": ".1.3.6.1.2.1.2.2.1.8",
            "ifLastChange": ".1.3.6.1.2.1.2.2.1.9",
            "ifInOctets": ".1.3.6.1.2.1.2.2.1.10",
            "ifInUcastPkts": ".1.3.6.1.2.1.2.2.1.11",
            "ifInNUcastPkts": ".1.3.6.1.2.1.2.2.1.12",
            "ifInDiscards": ".1.3.6.1.2.1.2.2.1.13",
            "ifInErrors": ".1.3.6.1.2.1.2.2.1.14",
            "ifInUnknownProtos": ".1.3.6.1.2.1.2.2.1.15",
            "ifOutOctets": ".1.3.6.1.2.1.2.2.1.16",
            "ifOutUcastPkts": ".1.3.6.1.2.1.2.2.1.17",
            "ifOutNUcastPkts": ".1.3.6.1.2.1.2.2.1.18",
            "ifOutDiscards": ".1.3.6.1.2.1.2.2.1.19",
            "ifOutErrors": ".1.3.6.1.2.1.2.2.1.20",
            "ifOutQLen": ".1.3.6.1.2.1.2.2.1.21",
            "ifSpecific": ".1.3.6.1.2.1.2.2.1.22",
        },
        'ESR': {
            'cpuUsage': '.1.3.6.1.4.1.15227.1.3.1.1.1',
            'memUsage': '.1.3.6.1.4.1.15227.1.3.1.1.2',
            'sessionNum': '.1.3.6.1.4.1.15227.1.3.1.1.3',
            'forwardRate': '.1.3.6.1.4.1.15227.1.3.1.1.5',
            'memTotal': '.1.3.6.1.4.1.15227.1.3.1.1.6',
            'memFree': '.1.3.6.1.4.1.15227.1.3.1.1.7',
            'powerState': '.1.3.6.1.4.1.15227.1.3.1.1.8',
            'CpuTemperature': '.1.3.6.1.4.1.15227.1.3.1.1.11'
        }
    }
    ips_by_device_type = {
        'OmniSwitch': ['10.10.10.68', '10.10.10.226', '10.10.10.227'],
        'ESR': ['10.10.10.56']
    }

    previous_data = {}

    while True:
        with ThreadPoolExecutor(max_workers=10) as executor:
            for device_type, oids in oids_by_device_type.items():
                for ip in ips_by_device_type[device_type]:
                    executor.submit(process_device, device_type, ip, oids, previous_data)
        time.sleep(5)

if __name__ == "__main__":
    main()